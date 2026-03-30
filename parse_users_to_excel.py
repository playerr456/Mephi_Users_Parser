#!/usr/bin/env python3
"""Export users from https://home.mephi.ru/users to an Excel file."""

from __future__ import annotations

import argparse
import getpass
import os
import re
import sys
from dataclasses import dataclass
from typing import Iterable, Optional
from urllib.parse import parse_qs, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

USERS_URL = "https://home.mephi.ru/users"
LOGIN_URL = "https://auth.mephi.ru/login"
NULL_GROUP = "NULL"

GROUP_PATTERNS = [
    re.compile(r"\b[A-ZА-ЯЁ][A-ZА-ЯЁ0-9]{0,7}-[A-ZА-ЯЁ0-9]{1,8}\b"),
]
WORD_PATTERN = re.compile(r"^[A-Za-zА-Яа-яЁё-]+$")


@dataclass(frozen=True)
class UserRecord:
    full_name: str
    group_number: str
    user_key: str = ""


def normalize_text(value: str) -> str:
    return " ".join(value.replace("\xa0", " ").split())


def looks_like_full_name(value: str) -> bool:
    value = normalize_text(value)
    if not value:
        return False

    words = [w.strip(".,;:()[]{}\"'") for w in value.split()]
    words = [w for w in words if w]
    if not (2 <= len(words) <= 4):
        return False

    if not all(WORD_PATTERN.match(word) for word in words):
        return False

    capitals = sum(1 for word in words if word[0].isalpha() and word[0].isupper())
    return capitals >= 2


def find_group_number(value: str) -> Optional[str]:
    text = normalize_text(value).upper().replace("—", "-").replace("–", "-")
    candidates: list[str] = []

    for pattern in GROUP_PATTERNS:
        for match in pattern.finditer(text):
            candidate = match.group(0)
            if any(char.isalpha() for char in candidate) and any(char.isdigit() for char in candidate):
                candidates.append(candidate)

    if not candidates:
        return None

    return max(candidates, key=len)


def infer_record(fragments: Iterable[str]) -> Optional[UserRecord]:
    parts = [normalize_text(part) for part in fragments if normalize_text(part)]
    if not parts:
        return None

    name: Optional[str] = None
    group: Optional[str] = None

    for part in parts:
        if not name and looks_like_full_name(part):
            name = part
        if not group:
            group = find_group_number(part)
        if name and group:
            break

    if not (name and group):
        combined = normalize_text(" ".join(parts))
        if not group:
            group = find_group_number(combined)

        if not name:
            for candidate in re.split(r"[|,;/]", combined):
                candidate = normalize_text(candidate)
                if looks_like_full_name(candidate):
                    name = candidate
                    break

    if name:
        return UserRecord(
            full_name=name,
            group_number=group or NULL_GROUP,
        )

    return None


def extract_from_table_rows(soup: BeautifulSoup) -> list[UserRecord]:
    records: list[UserRecord] = []

    for row in soup.select("table tr"):
        cells = [
            normalize_text(cell.get_text(" ", strip=True))
            for cell in row.find_all(["th", "td"])
        ]
        cells = [cell for cell in cells if cell]
        if len(cells) < 2:
            continue

        record = infer_record(cells)
        if record:
            records.append(record)

    return records


def extract_from_list_group_items(soup: BeautifulSoup) -> list[UserRecord]:
    records: list[UserRecord] = []

    for item in soup.select('a.list-group-item[href^="/users/"]'):
        inline = item.select_one(".inline-block") or item

        name: Optional[str] = None
        for piece in inline.stripped_strings:
            candidate = normalize_text(piece)
            if looks_like_full_name(candidate):
                name = candidate
                break

        if not name:
            continue

        group: Optional[str] = None
        for selector in (".text-sm", ".text-muted", "small"):
            for block in inline.select(selector):
                candidate = find_group_number(block.get_text(" ", strip=True))
                if candidate:
                    group = candidate
                    break
            if group:
                break

        if not group:
            for piece in inline.stripped_strings:
                candidate = normalize_text(piece)
                if candidate == name:
                    continue
                group = find_group_number(candidate)
                if group:
                    break

        href = item.get("href", "").strip()
        records.append(
            UserRecord(
                full_name=name,
                group_number=group or NULL_GROUP,
                user_key=href,
            )
        )

    return records


def extract_from_profile_links(soup: BeautifulSoup) -> list[UserRecord]:
    records: list[UserRecord] = []

    for link in soup.select("a[href]"):
        href = link.get("href", "")
        if "/users/" not in href:
            continue

        name = normalize_text(link.get_text(" ", strip=True))
        if not looks_like_full_name(name):
            continue

        context_text = ""
        for parent_tag in ("tr", "li", "article"):
            parent = link.find_parent(parent_tag)
            if parent:
                context_text = normalize_text(parent.get_text(" ", strip=True))
                if context_text:
                    break

        group = find_group_number(context_text) or NULL_GROUP
        records.append(UserRecord(full_name=name, group_number=group))

    return records


def extract_from_generic_blocks(soup: BeautifulSoup) -> list[UserRecord]:
    records: list[UserRecord] = []

    for block in soup.select("li, article, .card, .media, .user"):
        block_text = normalize_text(block.get_text(" ", strip=True))
        group = find_group_number(block_text)
        name: Optional[str] = None
        for piece in block.stripped_strings:
            candidate = normalize_text(piece)
            if looks_like_full_name(candidate):
                name = candidate
                break

        if name:
            records.append(
                UserRecord(
                    full_name=name,
                    group_number=group or NULL_GROUP,
                )
            )

    return records


def extract_users_from_page(soup: BeautifulSoup) -> list[UserRecord]:
    unique: dict[tuple[str, ...], UserRecord] = {}

    list_group_records = extract_from_list_group_items(soup)
    if list_group_records:
        for record in list_group_records:
            key = ("id", record.user_key) if record.user_key else (
                "name_group",
                record.full_name,
                record.group_number,
            )
            unique[key] = record
        return list(unique.values())

    for record in (
        extract_from_table_rows(soup)
        + extract_from_profile_links(soup)
        + extract_from_generic_blocks(soup)
    ):
        key = ("id", record.user_key) if record.user_key else (
            "name_group",
            record.full_name,
            record.group_number,
        )
        unique[key] = record

    return list(unique.values())


def page_number_from_url(url: str) -> Optional[int]:
    query = parse_qs(urlparse(url).query)
    for key in ("page", "p"):
        value = query.get(key)
        if not value:
            continue
        try:
            return int(value[0])
        except (TypeError, ValueError):
            continue
    return None


def find_next_page_url(soup: BeautifulSoup, current_url: str) -> Optional[str]:
    rel_next = soup.select_one("a[rel~=next][href]")
    if rel_next:
        return urljoin(current_url, rel_next["href"])

    current_page = page_number_from_url(current_url) or 1

    text_hints = {"next", "следующая", "далее", "вперед", ">", "»"}
    page_candidates: list[tuple[int, str]] = []

    for link in soup.select("a[href]"):
        href = link.get("href", "").strip()
        if not href or href == "#":
            continue

        absolute = urljoin(current_url, href)
        label = normalize_text(link.get_text(" ", strip=True)).lower()
        if label in text_hints:
            return absolute

        page_number = page_number_from_url(absolute)
        if page_number and page_number > current_page:
            page_candidates.append((page_number, absolute))

    if not page_candidates:
        return None

    page_candidates.sort(key=lambda item: item[0])
    return page_candidates[0][1]


def is_login_page(response: requests.Response) -> bool:
    lowered_url = response.url.lower()
    if "auth.mephi.ru/login" in lowered_url:
        return True

    content = response.text.lower()
    return 'id="login-form"' in content and 'name="username"' in content


def extract_login_error(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")
    for selector in (".alert-danger", ".alert", ".error", ".errors"):
        block = soup.select_one(selector)
        if block:
            message = normalize_text(block.get_text(" ", strip=True))
            if message:
                return message
    return None


def login(session: requests.Session, username: str, password: str, timeout: int) -> None:
    response = session.get(
        LOGIN_URL,
        params={"service": USERS_URL},
        timeout=timeout,
    )
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    form = soup.find("form", id="login-form") or soup.find("form")
    if not form:
        raise RuntimeError("Login form was not found. Website layout may have changed.")

    payload: dict[str, str] = {}
    for hidden in form.select("input[name]"):
        name = hidden.get("name")
        if not name:
            continue
        payload[name] = hidden.get("value", "")

    payload["username"] = username
    payload["password"] = password

    action = urljoin(response.url, form.get("action") or "/login")
    action_parts = urlparse(action)
    origin = f"{action_parts.scheme}://{action_parts.netloc}"
    submit_headers = {
        "Referer": response.url,
        "Origin": origin,
    }
    submit = session.post(
        action,
        data=payload,
        headers=submit_headers,
        timeout=timeout,
        allow_redirects=True,
    )
    if submit.status_code >= 500:
        submit.raise_for_status()
    if submit.status_code >= 400 and submit.status_code not in (401, 403):
        submit.raise_for_status()

    check = session.get(USERS_URL, timeout=timeout, allow_redirects=True)
    check.raise_for_status()
    if is_login_page(check):
        error_text = extract_login_error(check.text)
        if error_text:
            raise RuntimeError(f"Authentication failed: {error_text}")
        raise RuntimeError("Authentication failed. Check login/password.")


def collect_users(
    session: requests.Session,
    start_page: int,
    max_pages: int,
    timeout: int,
    verbose: bool,
) -> list[UserRecord]:
    if start_page <= 1:
        next_url: Optional[str] = USERS_URL
    else:
        next_url = f"{USERS_URL}?page={start_page}"
    visited: set[str] = set()
    unique: dict[tuple[str, ...], UserRecord] = {}

    page_index = 0
    while next_url and next_url not in visited and page_index < max_pages:
        page_index += 1
        visited.add(next_url)

        response = session.get(next_url, timeout=timeout, allow_redirects=True)
        response.raise_for_status()
        if is_login_page(response):
            raise RuntimeError("Session expired while loading users pages.")

        soup = BeautifulSoup(response.text, "html.parser")
        page_records = extract_users_from_page(soup)

        for record in page_records:
            key = ("id", record.user_key) if record.user_key else (
                "name_group",
                record.full_name,
                record.group_number,
            )
            unique[key] = record

        if verbose:
            page_label = page_number_from_url(response.url)
            page_text = str(page_label) if page_label is not None else "?"
            print(
                f"[page {page_text}] {response.url} -> "
                f"{len(page_records)} records, total {len(unique)}"
            )

        next_url = find_next_page_url(soup, response.url)

    records = list(unique.values())
    records.sort(key=lambda item: (item.full_name, item.group_number))
    return records


def write_excel(records: list[UserRecord], output_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "users"

    sheet.append(["ФИО", "Номер группы"])
    for record in records:
        sheet.append([record.full_name, record.group_number])

    for index in range(1, 3):
        letter = get_column_letter(index)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in sheet[letter]
        )
        sheet.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export users from home.mephi.ru/users to an Excel file.",
    )
    parser.add_argument(
        "-u",
        "--username",
        help="MEPhI login (or set MEPHI_USERNAME env var).",
    )
    parser.add_argument(
        "-p",
        "--password",
        help="MEPhI password (or set MEPHI_PASSWORD env var).",
    )
    parser.add_argument(
        "--creds-file",
        default="mephi_credentials.env",
        help=(
            "Path to credentials file (default: mephi_credentials.env). "
            "Supported keys: MEPHI_USERNAME/MEPHI_PASSWORD."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        default="mephi_users.xlsx",
        help="Output xlsx path (default: mephi_users.xlsx).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=30,
        help="HTTP timeout in seconds (default: 30).",
    )
    parser.add_argument(
        "--start-page",
        type=int,
        default=1,
        help="Start page number for parsing (default: 1).",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=9999,
        help="Maximum pages to scan (default: 9999).",
    )
    parser.add_argument(
        "--insecure",
        action="store_true",
        help="Disable TLS certificate verification.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print progress while parsing pages.",
    )
    return parser.parse_args()


def load_credentials_file(path: str) -> dict[str, str]:
    if not path or not os.path.exists(path):
        return {}

    credentials: dict[str, str] = {}
    with open(path, "r", encoding="utf-8") as file:
        for line_index, raw_line in enumerate(file, start=1):
            line = raw_line.lstrip("\ufeff").strip()
            if not line or line.startswith("#"):
                continue

            if "=" not in line:
                raise RuntimeError(
                    f"Invalid credentials file format at {path}:{line_index}. "
                    "Use KEY=VALUE."
                )

            key, value = line.split("=", 1)
            key = key.strip().upper()
            value = value.strip()
            value = value.strip("\"'")
            credentials[key] = value

    return credentials


def main() -> int:
    args = parse_args()

    file_credentials: dict[str, str] = {}
    try:
        file_credentials = load_credentials_file(args.creds_file)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    username = (
        args.username
        or os.getenv("MEPHI_USERNAME")
        or file_credentials.get("MEPHI_USERNAME")
        or file_credentials.get("USERNAME")
        or ""
    ).strip()
    if not username:
        username = input("MEPhI login: ").strip()

    password = (
        args.password
        or os.getenv("MEPHI_PASSWORD")
        or file_credentials.get("MEPHI_PASSWORD")
        or file_credentials.get("PASSWORD")
    )
    if not password:
        password = getpass.getpass("MEPhI password: ")

    if not username or not password:
        print("Login and password are required.", file=sys.stderr)
        return 1

    session = requests.Session()
    session.verify = not args.insecure
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/123.0 Safari/537.36"
            ),
        }
    )

    try:
        if args.start_page < 1:
            print("--start-page must be >= 1", file=sys.stderr)
            return 1

        login(session, username, password, timeout=args.timeout)
        users = collect_users(
            session,
            start_page=args.start_page,
            max_pages=args.max_pages,
            timeout=args.timeout,
            verbose=args.verbose,
        )
    except requests.RequestException as exc:
        print(f"Network error: {exc}", file=sys.stderr)
        return 2
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 3

    if not users:
        print(
            "No users were extracted. The page layout may have changed; "
            "inspect HTML selectors in the script.",
            file=sys.stderr,
        )
        return 4

    write_excel(users, args.output)
    print(f"Saved {len(users)} rows to: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
